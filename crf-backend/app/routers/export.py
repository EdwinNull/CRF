"""
数据导出路由
"""
from typing import Annotated, Optional, List
from datetime import datetime
from io import BytesIO

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.core.database import get_db
from app.models import User, Patient, Visit, AdverseEvent, ConcomitantMed
from app.routers.auth import get_current_active_user

router = APIRouter(prefix="/export")


def filter_patients(
    db: Session,
    current_user: User,
    center_ids: Optional[List[str]] = None,
    statuses: Optional[List[str]] = None,
    date_from: Optional[str] = None,
    date_to: Optional[str] = None
) -> List[Patient]:
    """筛选患者（多中心隔离 + 筛选条件）"""
    query = db.query(Patient)

    # 多中心隔离：仅 admin 可跨中心导出
    if current_user.role != "admin":
        query = query.filter(Patient.center_id == current_user.center_id)
    elif center_ids:
        query = query.filter(Patient.center_id.in_(center_ids))

    # 状态筛选
    if statuses:
        query = query.filter(Patient.status.in_(statuses))

    # 入组日期筛选
    if date_from:
        query = query.filter(Patient.enrollment_date >= date_from)
    if date_to:
        query = query.filter(Patient.enrollment_date <= date_to)

    patients = query.order_by(Patient.screening_no).all()
    return patients


def create_excel_full(patients: List[Patient], db: Session) -> BytesIO:
    """生成全量数据 Excel（4个Sheet）"""
    wb = Workbook()

    # Sheet 1: 患者基本信息
    ws_patients = wb.active
    ws_patients.title = "患者基本信息"
    headers_patients = [
        "筛选号", "随机号", "姓名缩写", "研究中心", "性别", "年龄",
        "身高(cm)", "体重(kg)", "入组日期", "患者状态", "退出原因", "退出日期"
    ]
    ws_patients.append(headers_patients)

    for row in ws_patients[1]:
        row.font = Font(bold=True)
        row.alignment = Alignment(horizontal="center")

    for p in patients:
        ws_patients.append([
            p.screening_no, p.randomization_no or "", p.name_abbr, p.center_id,
            p.gender, p.age, p.height or "", p.weight or "", p.enrollment_date,
            p.status, p.withdrawal_reason or "", p.withdrawal_date or ""
        ])

    # Sheet 2: 访视记录
    ws_visits = wb.create_sheet("访视记录")
    headers_visits = [
        "筛选号", "访视编号", "访视日期", "状态", "访视数据摘要"
    ]
    ws_visits.append(headers_visits)

    for row in ws_visits[1]:
        row.font = Font(bold=True)
        row.alignment = Alignment(horizontal="center")

    for p in patients:
        visits = db.query(Visit).filter(Visit.patient_id == p.id).order_by(Visit.visit_no).all()
        for v in visits:
            data_summary = f"{len(v.data)} 个字段" if v.data else "无数据"
            ws_visits.append([
                p.screening_no, v.visit_no, v.visit_date or "", v.status, data_summary
            ])

    # Sheet 3: 不良事件
    ws_ae = wb.create_sheet("不良事件")
    headers_ae = [
        "筛选号", "序号", "事件名称", "开始日期", "结束日期", "严重程度",
        "与研究药物关系", "转归", "是否SAE"
    ]
    ws_ae.append(headers_ae)

    for row in ws_ae[1]:
        row.font = Font(bold=True)
        row.alignment = Alignment(horizontal="center")

    severity_map = {1: "轻度", 2: "中度", 3: "重度"}
    relation_map = {1: "肯定有关", 2: "很可能有关", 3: "可能有关", 4: "可能无关", 5: "肯定无关"}
    outcome_map = {1: "无变化", 2: "病情恶化", 3: "恢复治愈", 4: "改善中", 5: "留有后遗症", 6: "死亡"}

    for p in patients:
        aes = db.query(AdverseEvent).filter(AdverseEvent.patient_id == p.id).order_by(AdverseEvent.seq_no).all()
        for ae in aes:
            ws_ae.append([
                p.screening_no, ae.seq_no, ae.event_name, ae.start_date, ae.end_date or "持续中",
                severity_map.get(ae.severity, ""), relation_map.get(ae.drug_relation, ""),
                outcome_map.get(ae.outcome, ""), "是" if ae.is_sae else "否"
            ])

    # Sheet 4: 合并用药
    ws_med = wb.create_sheet("合并用药")
    headers_med = [
        "筛选号", "序号", "药物名称", "适应症", "剂型", "剂量",
        "开始日期", "结束日期", "与研究药物关系"
    ]
    ws_med.append(headers_med)

    for row in ws_med[1]:
        row.font = Font(bold=True)
        row.alignment = Alignment(horizontal="center")

    for p in patients:
        meds = db.query(ConcomitantMed).filter(ConcomitantMed.patient_id == p.id).order_by(ConcomitantMed.seq_no).all()
        for med in meds:
            ws_med.append([
                p.screening_no, med.seq_no, med.drug_name, med.indication or "",
                med.dosage_form or "", med.dosage_amount or "", med.start_date,
                med.end_date or "持续中", med.drug_relation or ""
            ])

    # 保存到内存流
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


def create_excel_safety(patients: List[Patient], db: Session) -> BytesIO:
    """生成安全性数据 Excel（2个Sheet：患者基本信息 + 不良事件）"""
    wb = Workbook()

    # Sheet 1: 患者基本信息
    ws_patients = wb.active
    ws_patients.title = "患者基本信息"
    headers_patients = [
        "筛选号", "随机号", "姓名缩写", "研究中心", "性别", "年龄", "患者状态"
    ]
    ws_patients.append(headers_patients)

    for row in ws_patients[1]:
        row.font = Font(bold=True)
        row.alignment = Alignment(horizontal="center")

    for p in patients:
        ws_patients.append([
            p.screening_no, p.randomization_no or "", p.name_abbr, p.center_id,
            p.gender, p.age, p.status
        ])

    # Sheet 2: 不良事件
    ws_ae = wb.create_sheet("不良事件")
    headers_ae = [
        "筛选号", "序号", "事件名称", "描述", "开始日期", "结束日期",
        "严重程度", "与研究药物关系", "研究药物采取的措施", "其他处理措施",
        "转归", "是否SAE", "SAE类型"
    ]
    ws_ae.append(headers_ae)

    for row in ws_ae[1]:
        row.font = Font(bold=True)
        row.alignment = Alignment(horizontal="center")

    severity_map = {1: "轻度", 2: "中度", 3: "重度"}
    relation_map = {1: "肯定有关", 2: "很可能有关", 3: "可能有关", 4: "可能无关", 5: "肯定无关"}
    drug_measure_map = {1: "维持原状", 2: "暂停用药", 3: "减少剂量", 4: "停止用药", 5: "其他"}
    other_measure_map = {1: "无需特殊处理", 2: "给予对症治疗", 3: "住院治疗", 4: "其他"}
    outcome_map = {1: "无变化", 2: "病情恶化", 3: "恢复治愈", 4: "改善中", 5: "留有后遗症", 6: "死亡"}
    sae_type_map = {1: "死亡", 2: "危及生命", 3: "永久性伤残", 4: "住院或延长住院", 5: "先天性异常", 6: "其他重要医学事件", 7: "需要医学干预"}

    for p in patients:
        aes = db.query(AdverseEvent).filter(AdverseEvent.patient_id == p.id).order_by(AdverseEvent.seq_no).all()
        for ae in aes:
            ws_ae.append([
                p.screening_no, ae.seq_no, ae.event_name, ae.description or "",
                ae.start_date, ae.end_date or "持续中",
                severity_map.get(ae.severity, ""), relation_map.get(ae.drug_relation, ""),
                drug_measure_map.get(ae.drug_measure, ""), other_measure_map.get(ae.other_measure, ""),
                outcome_map.get(ae.outcome, ""), "是" if ae.is_sae else "否",
                sae_type_map.get(ae.sae_type, "") if ae.is_sae else ""
            ])

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output


@router.get("")
async def export_data(
    current_user: Annotated[User, Depends(get_current_active_user)],
    db: Session = Depends(get_db),
    mode: str = Query("full", pattern="^(full|safety)$", description="导出模式：full=全量数据，safety=安全性数据"),
    centers: Optional[str] = Query(None, description="中心ID列表（逗号分隔，如：01,02，仅admin可用）"),
    statuses: Optional[str] = Query(None, description="患者状态列表（逗号分隔，如：treatment,followup）"),
    date_from: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$", description="入组日期起始"),
    date_to: Optional[str] = Query(None, pattern=r"^\d{4}-\d{2}-\d{2}$", description="入组日期截止")
):
    """
    数据导出 API
    - 多中心隔离：doctor 仅能导出本中心，admin 可跨中心
    - 筛选条件：中心、状态、入组日期范围
    - 导出模式：full=全量数据（4个Sheet），safety=安全性数据（2个Sheet）
    """
    # 解析筛选参数
    center_ids = centers.split(",") if centers else None
    status_list = statuses.split(",") if statuses else None

    # 筛选患者
    patients = filter_patients(db, current_user, center_ids, status_list, date_from, date_to)

    if not patients:
        raise HTTPException(status_code=404, detail="未找到符合条件的患者数据")

    # 生成 Excel
    if mode == "full":
        excel_file = create_excel_full(patients, db)
        filename = f"CRF_全量数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    else:
        excel_file = create_excel_safety(patients, db)
        filename = f"CRF_安全性数据_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"

    # 返回文件流（中文文件名需用 RFC 5987 filename* 编码，否则 latin-1 编码报错）
    from urllib.parse import quote

    filename_ascii = "CRF_export.xlsx"
    headers = {
        "Content-Disposition": (
            f"attachment; filename=\"{filename_ascii}\"; "
            f"filename*=UTF-8''{quote(filename)}"
        )
    }

    return StreamingResponse(
        excel_file,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers
    )
